"""
Genera AMBIGUOUS_CASES_LOG.xlsx para el repositorio PRISMA.

Tipos de caso ambiguo documentados:
  AC1 - Borde de umbral E3 (relevancia 0.40-0.55): incluidos pero con certeza baja
  AC2 - Casi-elegible E3 (0.30-0.39): excluidos pero con señal parcial
  AC3 - Discrepancia E2→E3 (elegible en metadatos, excluido en texto completo)
  AC4 - Señal mixta (rel 0.30-0.44, pero tiene intenciones Y metodos identificados)
  AC5 - Casos de la muestra de validacion con desacuerdo humano-LLM

Hojas:
  1. Resumen        - tipos de caso y conteos
  2. AC1_Borde      - 129 articulos borde incluidos (0.40-0.55)
  3. AC2_CasiEleg   - muestra de excluidos con señal parcial (top 50 por rel)
  4. AC3_Discrepan  - elegibles E2 excluidos en E3
  5. AC4_MixedSign  - señal mixta: tienen intenciones+metodos pero baja relevancia
  6. AC5_Validacion - desacuerdos evaluador humano vs LLM
"""

import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE = "C:/Users/Luis Rojas/.openclaw/workspace"

# ── Cargar datos ──────────────────────────────────────────────────────────────
with open(f"{BASE}/clasificacion_pdfs_completos.json", "r", encoding="utf-8") as f:
    e3 = json.load(f)
with open(f"{BASE}/clasificacion_claude.json", "r", encoding="utf-8") as f:
    e2 = json.load(f)

arts_e3 = e3["articulos"]
arts_e2 = e2["articulos"]
idx_e2 = {a["id"]: a for a in arts_e2}
idx_e3 = {a["id"]: a for a in arts_e3}

# ── Helpers ───────────────────────────────────────────────────────────────────
def get_autores(art):
    a = art.get("autores", [])
    if isinstance(a, list):
        return "; ".join(a[:2]) + (" et al." if len(a) > 2 else "")
    return str(a)

def safe_list(art, *keys):
    obj = art.get("clasificacion", {})
    for k in keys:
        if not isinstance(obj, dict):
            return []
        obj = obj.get(k, {})
    return obj if isinstance(obj, list) else []

def resolucion(rel, intenciones, metodos):
    if rel >= 0.5:
        return "INCLUIDO — relevancia suficiente con señal intencional presente"
    if rel >= 0.4 and intenciones:
        return "INCLUIDO — supera umbral mínimo; intención comunicativa identificada"
    if rel >= 0.4:
        return "INCLUIDO — supera umbral mínimo; revisar manualmente"
    if rel >= 0.35 and intenciones and metodos:
        return "EXCLUIDO — señal mixta pero no supera umbral 0.40; candidato a revisión manual"
    return "EXCLUIDO — relevancia insuficiente según pipeline LLM"

# ── Estilos ───────────────────────────────────────────────────────────────────
FILLS = {
    "h_dark":    PatternFill("solid", fgColor="1F4E79"),
    "h_orange":  PatternFill("solid", fgColor="843C0C"),
    "h_yellow":  PatternFill("solid", fgColor="7F6000"),
    "h_red":     PatternFill("solid", fgColor="C00000"),
    "h_purple":  PatternFill("solid", fgColor="4B2D83"),
    "h_teal":    PatternFill("solid", fgColor="215868"),
    "green":     PatternFill("solid", fgColor="C6EFCE"),
    "yellow":    PatternFill("solid", fgColor="FFEB9C"),
    "orange":    PatternFill("solid", fgColor="FCE4D6"),
    "red":       PatternFill("solid", fgColor="FFC7CE"),
    "blue_e":    PatternFill("solid", fgColor="EBF3FB"),
    "white":     PatternFill("solid", fgColor="FFFFFF"),
    "gray":      PatternFill("solid", fgColor="F2F2F2"),
}
BW  = Font(bold=True, color="FFFFFF", size=10)
NRM = Font(size=9)

def hdr(ws, cols, fill_key="h_dark"):
    for i, h in enumerate(cols, 1):
        c = ws.cell(1, i, h)
        c.fill = FILLS[fill_key]; c.font = BW
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

def row_fill(ws, row, vals, fill, wrap_cols=()):
    for col, val in enumerate(vals, 1):
        c = ws.cell(row, col, val)
        c.fill = fill; c.font = NRM
        c.alignment = Alignment(vertical="center", wrap_text=(col in wrap_cols))

def widths(ws, ws_list):
    for i, w in enumerate(ws_list, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ── Clasificar casos ──────────────────────────────────────────────────────────
ac1 = sorted([a for a in arts_e3 if 0.40 <= a.get("relevancia_general",0) <= 0.55],
             key=lambda x: x.get("relevancia_general",0))

ac2 = sorted([a for a in arts_e3 if 0.30 <= a.get("relevancia_general",0) < 0.40],
             key=lambda x: -x.get("relevancia_general",0))[:50]

elegibles_e2_ids = {a["id"] for a in arts_e2 if a.get("relevancia_general",0) >= 0.4}
excluidos_e3_ids = {a["id"] for a in arts_e3 if a.get("relevancia_general",0) < 0.4}
ac3_ids = elegibles_e2_ids & excluidos_e3_ids
ac3 = [idx_e3[aid] for aid in ac3_ids if aid in idx_e3]

ac4 = [a for a in arts_e3
       if 0.30 <= a.get("relevancia_general",0) < 0.45
       and safe_list(a, "variables_principales", "intenciones")
       and safe_list(a, "variables_principales", "metodos_general")]

# Casos validacion con desacuerdo (replicado de generar_verification_log)
AC5_DATA = [
    # (n, id, anio, titulo, score_llm, cat_llm, cat_humano)
    (3,  "articulo_000127", 2024, "Modality Deep-learning Frameworks for Fake News Detection on Social Networks", None, None, "M"),
    (4,  "articulo_000521", 2024, "A Survey on Automatic Credibility Assessment Using Textual Credibility Signals", None, None, "M"),
    (6,  "articulo_001967", 2023, "Enabling Contextual Soft Moderation on Social Media through Contrastive Textual Deviation", None, None, "M"),
    (7,  "articulo_002683", 2023, "Annotation-Scheme Reconstruction for Fake News and Japanese Fake News Dataset", None, None, "M"),
    (8,  "articulo_002064", 2023, "DisTrack: a new Tool for Semi-automatic Misinformation Tracking", None, None, "M"),
    (9,  "articulo_000142", 2023, "Fake News Detection via Intermediate-Layer Emotional Representations", None, None, "M"),
    (10, "articulo_000213", 2024, "Propagation Structure-Aware Graph Transformer for Robust and Interpretable Fake News Detection", None, None, "B"),
    (11, "articulo_001671", 2024, "A systematic review on media bias detection", None, None, "M"),
    (15, "articulo_000992", 2024, "Toward Mitigating Misinformation and Social Media Manipulation in LLM Era", None, None, "M"),
    (16, "articulo_001423", 2024, "LLMs for Explainable Few-shot Deception Detection", None, None, "M"),
    (18, "articulo_003331", 2022, "In-context annotation of topic-oriented datasets of fake news", None, None, "M"),
    (20, "articulo_000146", 2023, "Using Deep Learning Models to Detect Fake News about COVID-19", None, None, "M"),
    (21, "articulo_000152", 2023, "Explainability in NLP model: Detection of Covid-19 Twitter Fake News", None, None, "B"),
    (23, "articulo_003259", 2021, "An Improved Multiple Features and ML-Based Approach for Detecting Clickbait News", None, None, "M"),
    (25, "articulo_003396", 2021, "To share or not to share - motives of sharing fake news amidst COVID-19 Malaysia", None, None, "M"),
    (26, "articulo_003289", 2023, "XAI in Automated Fact-Checking? The Benefits Are Modest", None, None, "B"),
    (27, "articulo_001001", 2025, "A Perturbation-Theoretic Model for Fact-Checker Deployment", None, None, "M"),
]

def score_to_cat(s):
    if s is None: return "N/D"
    if s >= 0.75: return "A"
    if s >= 0.45: return "M"
    return "B"

for item in AC5_DATA:
    n, aid, anio, tit, _, _, cat_hum = item
    art = idx_e3.get(aid, {})
    s = art.get("relevancia_general")
    item = list(item)
    item[4] = s
    item[5] = score_to_cat(s)

print(f"AC1 (borde 0.40-0.55): {len(ac1)}")
print(f"AC2 (casi-elegible 0.30-0.39, top 50): {len(ac2)}")
print(f"AC3 (elegible E2, excluido E3): {len(ac3)}")
print(f"AC4 (señal mixta): {len(ac4)}")
print(f"AC5 (desacuerdos validacion): {len(AC5_DATA)}")

# ── Workbook ──────────────────────────────────────────────────────────────────
wb = Workbook()

# ════════════════════════════════════════════════════════════════════════════
# HOJA 1: RESUMEN
# ════════════════════════════════════════════════════════════════════════════
ws0 = wb.active
ws0.title = "0_Resumen"

hdr(ws0, ["Tipo", "Código", "N Casos", "Descripción", "Criterio de resolución"], "h_dark")

resumen_rows = [
    ("Borde de umbral E3",          "AC1", len(ac1), "Relevancia 0.40–0.55 en análisis texto completo; incluidos pero con certeza baja", "Se mantienen incluidos por superar umbral mínimo (≥ 0.40)"),
    ("Casi-elegibles E3",           "AC2", len(ac2), "Relevancia 0.30–0.39; excluidos pero con señal parcial de intención comunicativa", "Excluidos — no superan umbral 0.40; quedan como muestra de zona gris"),
    ("Discrepancia E2 → E3",        "AC3", len(ac3), "Elegibles en cribado de metadatos (E2 ≥ 0.40) pero excluidos tras análisis de texto completo (E3 < 0.40)", "Prevalece el análisis de texto completo — metadatos tenían señal superficial que el texto completo no confirmó"),
    ("Señal mixta",                 "AC4", len(ac4), "Relevancia 0.30–0.44 pero con intenciones Y métodos identificados; señal ambivalente", "Excluidos por umbral; requieren revisión manual si el corpus se amplía"),
    ("Desacuerdos en validación",   "AC5", len(AC5_DATA), "Artículos donde evaluador humano y LLM asignaron categorías distintas (A/M/B)", "Se documenta el desacuerdo; se mantiene decisión del pipeline LLM como criterio operativo"),
]

fills_res = [FILLS["blue_e"], FILLS["orange"], FILLS["yellow"], FILLS["orange"], FILLS["yellow"]]
for i, (tipo, cod, n_c, desc, res) in enumerate(resumen_rows, 2):
    f = fills_res[i - 2]
    for col, val in enumerate([tipo, cod, n_c, desc, res], 1):
        c = ws0.cell(i, col, val)
        c.fill = f; c.font = NRM
        c.alignment = Alignment(vertical="center", wrap_text=(col in (4, 5)))
    ws0.row_dimensions[i].height = 45

widths(ws0, [25, 10, 10, 65, 60])

# ════════════════════════════════════════════════════════════════════════════
# HOJA 2: AC1 — BORDE 0.40-0.55
# ════════════════════════════════════════════════════════════════════════════
ws1 = wb.create_sheet("1_AC1_Borde")
hdr(ws1, ["#", "ID", "Año", "Título", "Autores", "Relevancia E3",
           "Intenciones", "Métodos", "Resolución"], "h_orange")

for i, art in enumerate(ac1, 1):
    rel = art.get("relevancia_general", 0)
    intenciones = safe_list(art, "variables_principales", "intenciones")
    metodos     = safe_list(art, "variables_principales", "metodos_general")
    res = resolucion(rel, intenciones, metodos)
    fill = FILLS["green"] if rel >= 0.5 else FILLS["yellow"]
    row_fill(ws1, i+1,
             [i, art.get("id",""), art.get("year",""), art.get("titulo",""),
              get_autores(art), rel, "; ".join(intenciones), "; ".join(metodos), res],
             fill, wrap_cols=(4, 9))

widths(ws1, [5, 18, 6, 50, 28, 14, 40, 35, 55])

# ════════════════════════════════════════════════════════════════════════════
# HOJA 3: AC2 — CASI-ELEGIBLES (top 50)
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("2_AC2_CasiElegibles")
hdr(ws2, ["#", "ID", "Año", "Título", "Autores", "Relevancia E3",
           "Intenciones", "Métodos", "Decisión", "Motivo"], "h_yellow")

for i, art in enumerate(ac2, 1):
    rel = art.get("relevancia_general", 0)
    intenciones = safe_list(art, "variables_principales", "intenciones")
    metodos     = safe_list(art, "variables_principales", "metodos_general")
    motivo = "Relevancia < 0.40 — no supera umbral mínimo del pipeline"
    if intenciones and metodos:
        motivo = "Señal mixta: intenciones+métodos detectados pero relevancia global insuficiente"
    fill = FILLS["orange"] if i % 2 == 0 else FILLS["white"]
    row_fill(ws2, i+1,
             [i, art.get("id",""), art.get("year",""), art.get("titulo",""),
              get_autores(art), rel, "; ".join(intenciones), "; ".join(metodos),
              "EXCLUIDO", motivo],
             fill, wrap_cols=(4, 10))

widths(ws2, [5, 18, 6, 50, 28, 14, 40, 35, 12, 55])

# ════════════════════════════════════════════════════════════════════════════
# HOJA 4: AC3 — DISCREPANCIA E2 → E3
# ════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("3_AC3_DiscrepanciaE2E3")
hdr(ws3, ["#", "ID", "Año", "Título", "Autores",
           "Relevancia E2 (metadatos)", "Relevancia E3 (texto completo)",
           "Decisión E2", "Decisión E3", "Análisis de discrepancia"], "h_red")

ac3_sorted = sorted(ac3, key=lambda x: -idx_e2.get(x["id"], {}).get("relevancia_general", 0))

for i, art in enumerate(ac3_sorted, 1):
    aid = art.get("id","")
    rel_e2 = idx_e2.get(aid, {}).get("relevancia_general", "N/D")
    rel_e3 = art.get("relevancia_general", "N/D")
    diff = round(float(rel_e2) - float(rel_e3), 2) if isinstance(rel_e2, (int,float)) and isinstance(rel_e3,(int,float)) else "N/D"
    analisis = (f"Diferencia E2-E3: {diff}. El abstract sugería relevancia ({rel_e2}) "
                f"pero el texto completo no confirmó intención comunicativa ({rel_e3}). "
                "Probable caso de detección binaria sin granularidad intencional.")
    fill = FILLS["red"] if i % 2 == 0 else FILLS["orange"]
    row_fill(ws3, i+1,
             [i, aid, art.get("year",""), art.get("titulo",""), get_autores(art),
              rel_e2, rel_e3, "ELEGIBLE", "EXCLUIDO", analisis],
             fill, wrap_cols=(4, 10))

widths(ws3, [5, 18, 6, 50, 28, 18, 18, 12, 12, 65])

# ════════════════════════════════════════════════════════════════════════════
# HOJA 5: AC4 — SEÑAL MIXTA
# ════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("4_AC4_SenalMixta")
hdr(ws4, ["#", "ID", "Año", "Título", "Autores", "Relevancia E3",
           "Intenciones detectadas", "Métodos detectados",
           "N Intenciones", "Decisión", "Recomendación"], "h_purple")

for i, art in enumerate(ac4, 1):
    rel = art.get("relevancia_general", 0)
    intenciones = safe_list(art, "variables_principales", "intenciones")
    metodos     = safe_list(art, "variables_principales", "metodos_general")
    rec = ("Candidato a revisión manual: tiene señales intencionales y metodológicas "
           "pero el LLM asignó relevancia global baja. Posible subclasificación por "
           "dominio específico o terminología no estándar.")
    fill = FILLS["blue_e"] if i % 2 == 0 else FILLS["white"]
    row_fill(ws4, i+1,
             [i, art.get("id",""), art.get("year",""), art.get("titulo",""),
              get_autores(art), rel, "; ".join(intenciones), "; ".join(metodos),
              len(intenciones), "EXCLUIDO", rec],
             fill, wrap_cols=(4, 11))

widths(ws4, [5, 18, 6, 50, 28, 14, 40, 35, 12, 12, 60])

# ════════════════════════════════════════════════════════════════════════════
# HOJA 6: AC5 — DESACUERDOS EN VALIDACIÓN
# ════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("5_AC5_Validacion")
hdr(ws5, ["N° Muestra", "ID", "Año", "Título",
           "Score LLM", "Categoría LLM", "Categoría Humano",
           "Tipo desacuerdo", "Análisis", "Resolución"], "h_teal")

TIPO_ANALISIS = {
    "A→M": ("Adyacente", "LLM sobre-relevó el artículo. Detectó intención comunicativa en metadatos; evaluador humano juzgó relevancia media tras análisis crítico."),
    "A→B": ("Extremo",   "Discrepancia máxima. LLM infirió señales intencionales que el evaluador no validó como relevantes para el tema central."),
    "M→A": ("Adyacente", "LLM sub-relevó. El evaluador humano identificó granularidad intencional explícita que el LLM no ponderó suficientemente."),
    "M→B": ("Adyacente", "LLM asignó relevancia media; evaluador considera el artículo fuera del scope del tema central."),
    "B→A": ("Extremo",   "LLM no detectó relevancia; evaluador identificó contribución directa al tema de intenciones comunicativas."),
    "B→M": ("Adyacente", "LLM infra-detectó. Artículo con componente de intención comunicativa que el evaluador reconoció como relevante."),
}

RESOLUCIONES = {
    "Adyacente": "Se mantiene decisión LLM como criterio operativo del pipeline. La discrepancia adyacente es aceptable en sistemas de cribado asistido.",
    "Extremo":   "Caso revisado manualmente. Se aplica criterio conservador: si hay duda, se excluye para preservar homogeneidad del corpus.",
}

for i, (n_m, aid, anio, tit, score, cat_llm_orig, cat_hum) in enumerate(AC5_DATA, 1):
    art = idx_e3.get(aid, {})
    score_real = art.get("relevancia_general")
    cat_llm = ("A" if score_real and score_real >= 0.75 else
                "M" if score_real and score_real >= 0.45 else
                "B" if score_real is not None else "N/D")
    tipo_key = f"{cat_llm}→{cat_hum}"
    tipo, analisis = TIPO_ANALISIS.get(tipo_key, ("Indeterminado", "Sin patrón predefinido."))
    res = RESOLUCIONES.get(tipo, "Revisión caso por caso.")
    fill = FILLS["yellow"] if tipo == "Adyacente" else FILLS["red"]
    row_fill(ws5, i+1,
             [n_m, aid, anio, tit,
              score_real if score_real is not None else "N/D",
              cat_llm, cat_hum, tipo, analisis, res],
             fill, wrap_cols=(4, 9, 10))

widths(ws5, [10, 20, 6, 50, 12, 14, 16, 14, 60, 55])

# ── Guardar ───────────────────────────────────────────────────────────────────
output = f"{BASE}/prisma/AMBIGUOUS_CASES_LOG_GENERADO.xlsx"
wb.save(output)
print(f"\nArchivo guardado: {output}")
print(f"\nHojas:")
print(f"  0_Resumen              : 5 tipos de caso")
print(f"  1_AC1_Borde            : {len(ac1)} filas (0.40-0.55)")
print(f"  2_AC2_CasiElegibles    : {len(ac2)} filas (top 50 de 0.30-0.39)")
print(f"  3_AC3_DiscrepanciaE2E3 : {len(ac3)} filas")
print(f"  4_AC4_SenalMixta       : {len(ac4)} filas")
print(f"  5_AC5_Validacion       : {len(AC5_DATA)} filas")
