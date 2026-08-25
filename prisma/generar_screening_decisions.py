"""
Genera SCREENING_DECISIONS.xlsx para el repositorio PRISMA.

Fuentes:
  - clasificacion_claude.json   -> Etapa 2 (cribado metadatos, n=3575)
  - clasificacion_pdfs_completos.json -> Etapa 3 (texto completo, n=549)

Logica de decision:
  Etapa 2: relevancia >= 0.4 AND tiene al menos 1 intencion AND tiene al menos 1 metodo -> ELEGIBLE
           cualquier condicion que falle -> EXCLUIDO con motivo
  Etapa 3: relevancia >= 0.4 -> INCLUIDO; < 0.4 -> EXCLUIDO
           articulos elegibles en Etapa 2 sin PDF -> EXCLUIDO (CE5 - texto no disponible)
"""

import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = "C:/Users/Luis Rojas/.openclaw/workspace"

# ── Cargar datos ──────────────────────────────────────────────────────────────
with open(f"{BASE}/clasificacion_claude.json", "r", encoding="utf-8") as f:
    claude_data = json.load(f)

with open(f"{BASE}/clasificacion_pdfs_completos.json", "r", encoding="utf-8") as f:
    etapa3_data = json.load(f)

arts_e2 = claude_data["articulos"]          # 3575
arts_e3 = etapa3_data["articulos"]          # 549

# Índice Etapa 3 por ID
idx_e3 = {a["id"]: a for a in arts_e3}

# IDs con PDF analizado
ids_con_pdf = set(idx_e3.keys())

# ── Helpers ───────────────────────────────────────────────────────────────────
def get_intenciones(art):
    try:
        return art["clasificacion"]["variables_principales"]["intenciones"]
    except (KeyError, TypeError):
        return []

def get_metodos(art):
    try:
        return art["clasificacion"]["variables_principales"]["metodos_general"]
    except (KeyError, TypeError):
        return []

def get_autores(art):
    auths = art.get("autores", [])
    if isinstance(auths, list):
        return "; ".join(auths[:3]) + (" et al." if len(auths) > 3 else "")
    return str(auths)

def motivo_exclusion_e2(art):
    rel = art.get("relevancia_general", 0)
    intenciones = get_intenciones(art)
    metodos = get_metodos(art)
    motivos = []
    if rel < 0.4:
        motivos.append(f"Relevancia insuficiente ({rel})")
    if not intenciones:
        motivos.append("Sin intención comunicativa identificada")
    if not metodos:
        motivos.append("Sin método NLP/ML identificado")
    if not motivos:
        # Puede ser excluido por periodo
        year = art.get("year")
        if year and (int(year) < 2020 or int(year) > 2026):
            motivos.append(f"Fuera del período 2020-2026 (año: {year})")
    return "; ".join(motivos) if motivos else "CE1 - No cumple criterios combinados"

def decision_e2(art):
    rel = art.get("relevancia_general", 0)
    intenciones = get_intenciones(art)
    metodos = get_metodos(art)
    year = art.get("year")
    en_periodo = True
    if year:
        try:
            en_periodo = 2020 <= int(year) <= 2026
        except ValueError:
            en_periodo = True
    if rel >= 0.4 and intenciones and metodos and en_periodo:
        return "ELEGIBLE"
    return "EXCLUIDO"

# ── Clasificar Etapa 2 ────────────────────────────────────────────────────────
elegibles_e2 = []
excluidos_e2 = []

for art in arts_e2:
    dec = decision_e2(art)
    if dec == "ELEGIBLE":
        elegibles_e2.append(art)
    else:
        excluidos_e2.append(art)

# Artículos elegibles sin PDF (excluidos CE5 en Etapa 3)
sin_pdf = [a for a in elegibles_e2 if a["id"] not in ids_con_pdf]

print(f"Etapa 2: {len(arts_e2)} totales | {len(elegibles_e2)} elegibles | {len(excluidos_e2)} excluidos")
print(f"Etapa 3: {len(arts_e3)} con PDF | {len(sin_pdf)} sin PDF (CE5)")

# ── Crear Workbook ────────────────────────────────────────────────────────────
wb = Workbook()

# ════════════════════════════════════════════════════════════════════════════
# HOJA 1: RESUMEN
# ════════════════════════════════════════════════════════════════════════════
ws_res = wb.active
ws_res.title = "Resumen PRISMA"

encabezado_fill = PatternFill("solid", fgColor="1F4E79")
subenc_fill     = PatternFill("solid", fgColor="2E75B6")
verde_fill      = PatternFill("solid", fgColor="C6EFCE")
rojo_fill       = PatternFill("solid", fgColor="FFC7CE")
amarillo_fill   = PatternFill("solid", fgColor="FFEB9C")
blanco_fill     = PatternFill("solid", fgColor="FFFFFF")

bold_white = Font(bold=True, color="FFFFFF", size=12)
bold_dark  = Font(bold=True, color="1F4E79", size=11)
normal     = Font(size=10)

def cell_style(ws, row, col, value, fill=None, font=None, align="left", wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    if fill:
        c.fill = fill
    if font:
        c.font = font
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    return c

resumen_data = [
    ("PROCESO DE SELECCIÓN PRISMA 2020", None, None),
    ("Revisión: Identificación Granular de Intenciones Comunicativas", None, None),
    ("Autor: Luis Rojas Rubio | Doctorado en Informática - UDAEU", None, None),
    ("", None, None),
    ("ETAPA", "N", "Descripción"),
    ("Registros identificados (6 BD)", len(arts_e2), "Scopus, WoS, ACM, IEEE Xplore, Scholar, SciELO"),
    ("Excluidos en Etapa 2", len(excluidos_e2), "Relevancia < 0.4 o sin intención/método identificado"),
    ("Elegibles tras Etapa 2", len(elegibles_e2), "Cribado semántico de metadatos (Claude Sonnet 5, ≥ 0.4)"),
    ("Sin texto completo (CE5)", len(sin_pdf), "Acceso restringido; sin respuesta de autores"),
    ("Con texto completo (Etapa 3)", len(arts_e3), "PDF obtenido y analizado"),
    ("Excluidos en Etapa 3", len([a for a in arts_e3 if a.get("relevancia_general", 0) < 0.4]), "Relevancia < 0.4 en análisis de texto completo"),
    ("INCLUIDOS en síntesis", len([a for a in arts_e3 if a.get("relevancia_general", 0) >= 0.4]), "Corpus final de síntesis cuantitativa"),
]

for i, (etapa, n, desc) in enumerate(resumen_data, start=1):
    if i == 1:
        c = cell_style(ws_res, i, 1, etapa, fill=encabezado_fill, font=bold_white, align="center")
        ws_res.merge_cells(f"A{i}:C{i}")
    elif i in (2, 3):
        c = cell_style(ws_res, i, 1, etapa, fill=subenc_fill, font=Font(color="FFFFFF", size=10))
        ws_res.merge_cells(f"A{i}:C{i}")
    elif i == 4:
        pass
    elif i == 5:
        for j, val in enumerate([etapa, n, desc], start=1):
            cell_style(ws_res, i, j, val, fill=subenc_fill, font=bold_white, align="center")
    else:
        fills = [blanco_fill, blanco_fill, blanco_fill]
        if n is not None:
            if "Excluidos" in etapa or "Sin texto" in etapa:
                fills = [rojo_fill, rojo_fill, rojo_fill]
            elif "INCLUIDOS" in etapa:
                fills = [verde_fill, verde_fill, verde_fill]
            elif "Elegibles" in etapa or "Con texto" in etapa:
                fills = [amarillo_fill, amarillo_fill, amarillo_fill]
        cell_style(ws_res, i, 1, etapa, fill=fills[0], font=bold_dark)
        cell_style(ws_res, i, 2, n, fill=fills[1], font=Font(bold=True, size=11), align="center")
        cell_style(ws_res, i, 3, desc, fill=fills[2], font=normal)

ws_res.column_dimensions["A"].width = 42
ws_res.column_dimensions["B"].width = 10
ws_res.column_dimensions["C"].width = 58

# ════════════════════════════════════════════════════════════════════════════
# HOJA 2: ETAPA 2 — ELEGIBLES
# ════════════════════════════════════════════════════════════════════════════
ws_e2i = wb.create_sheet("Etapa2_Elegibles")

headers_e2 = ["ID", "Título", "Autores", "Año", "DOI",
               "Relevancia LLM", "Decisión", "Intenciones", "Métodos", "Plataforma", "Idioma"]

for col, h in enumerate(headers_e2, start=1):
    c = ws_e2i.cell(row=1, column=col, value=h)
    c.fill = subenc_fill
    c.font = bold_white
    c.alignment = Alignment(horizontal="center", vertical="center")

for row, art in enumerate(elegibles_e2, start=2):
    intenciones = "; ".join(get_intenciones(art))
    metodos     = "; ".join(get_metodos(art))
    try:
        plataforma = "; ".join(art["clasificacion"]["variables_adicionales"].get("plataforma", []))
        idioma     = "; ".join(art["clasificacion"]["variables_adicionales"].get("linguistica", []))
    except (KeyError, TypeError):
        plataforma = ""
        idioma = ""

    vals = [
        art.get("id", ""),
        art.get("titulo", ""),
        get_autores(art),
        art.get("year", ""),
        art.get("doi", ""),
        art.get("relevancia_general", ""),
        "ELEGIBLE",
        intenciones,
        metodos,
        plataforma,
        idioma,
    ]
    fill = verde_fill if row % 2 == 0 else blanco_fill
    for col, val in enumerate(vals, start=1):
        c = ws_e2i.cell(row=row, column=col, value=val)
        c.fill = fill
        c.font = normal
        c.alignment = Alignment(wrap_text=(col == 2), vertical="center")

col_widths_e2 = [16, 55, 30, 6, 32, 14, 12, 40, 35, 20, 14]
for i, w in enumerate(col_widths_e2, start=1):
    ws_e2i.column_dimensions[get_column_letter(i)].width = w

ws_e2i.freeze_panes = "A2"

# ════════════════════════════════════════════════════════════════════════════
# HOJA 3: ETAPA 2 — EXCLUIDOS
# ════════════════════════════════════════════════════════════════════════════
ws_e2x = wb.create_sheet("Etapa2_Excluidos")

headers_e2x = ["ID", "Título", "Autores", "Año", "DOI",
                "Relevancia LLM", "Decisión", "Motivo Exclusión", "Criterio"]

for col, h in enumerate(headers_e2x, start=1):
    c = ws_e2x.cell(row=1, column=col, value=h)
    c.fill = PatternFill("solid", fgColor="C00000")
    c.font = bold_white
    c.alignment = Alignment(horizontal="center", vertical="center")

for row, art in enumerate(excluidos_e2, start=2):
    rel = art.get("relevancia_general", 0)
    intenciones = get_intenciones(art)
    metodos = get_metodos(art)
    motivo = motivo_exclusion_e2(art)

    # Asignar criterio SPIDER
    criterio = "CE1"
    if not intenciones and not metodos:
        criterio = "CE1"
    elif rel < 0.4:
        criterio = "CE1"
    else:
        criterio = "CE2"

    vals = [
        art.get("id", ""),
        art.get("titulo", ""),
        get_autores(art),
        art.get("year", ""),
        art.get("doi", ""),
        rel,
        "EXCLUIDO",
        motivo,
        criterio,
    ]
    fill = PatternFill("solid", fgColor="FFF2CC") if row % 2 == 0 else blanco_fill
    for col, val in enumerate(vals, start=1):
        c = ws_e2x.cell(row=row, column=col, value=val)
        c.fill = fill
        c.font = normal
        c.alignment = Alignment(wrap_text=(col in (2, 8)), vertical="center")

col_widths_e2x = [16, 55, 30, 6, 32, 14, 12, 50, 12]
for i, w in enumerate(col_widths_e2x, start=1):
    ws_e2x.column_dimensions[get_column_letter(i)].width = w

ws_e2x.freeze_panes = "A2"

# ════════════════════════════════════════════════════════════════════════════
# HOJA 4: ETAPA 3 — INCLUIDOS (relevancia >= 0.4)
# ════════════════════════════════════════════════════════════════════════════
ws_e3i = wb.create_sheet("Etapa3_Incluidos")

headers_e3 = ["ID", "Título", "Autores", "Año",
               "Relevancia E3", "Decisión",
               "Intenciones", "Métodos", "Datasets", "Métricas",
               "Plataforma", "Idioma", "Hallazgos Emergentes (N)"]

for col, h in enumerate(headers_e3, start=1):
    c = ws_e3i.cell(row=1, column=col, value=h)
    c.fill = PatternFill("solid", fgColor="375623")
    c.font = bold_white
    c.alignment = Alignment(horizontal="center", vertical="center")

incluidos_e3 = [a for a in arts_e3 if a.get("relevancia_general", 0) >= 0.4]

for row, art in enumerate(incluidos_e3, start=2):
    try:
        intenciones = "; ".join(art["clasificacion"]["variables_principales"].get("intenciones", []))
        metodos     = "; ".join(art["clasificacion"]["variables_principales"].get("metodos_general", []))
        datasets    = "; ".join(art["clasificacion"]["variables_principales"].get("dataset", [])[:3])
        metricas    = "; ".join(art["clasificacion"]["variables_principales"].get("metricas", []))
        plataforma  = "; ".join(art["clasificacion"]["variables_adicionales"].get("plataforma", []))
        idioma      = "; ".join(art["clasificacion"]["variables_adicionales"].get("linguistica", []))
    except (KeyError, TypeError):
        intenciones = metodos = datasets = metricas = plataforma = idioma = ""

    hallazgos = art.get("hallazgos_nuevos", [])
    n_hallazgos = len(hallazgos) if isinstance(hallazgos, list) else 0

    vals = [
        art.get("id", ""),
        art.get("titulo", ""),
        get_autores(art),
        art.get("year", ""),
        art.get("relevancia_general", ""),
        "INCLUIDO",
        intenciones,
        metodos,
        datasets,
        metricas,
        plataforma,
        idioma,
        n_hallazgos,
    ]
    fill = verde_fill if row % 2 == 0 else blanco_fill
    for col, val in enumerate(vals, start=1):
        c = ws_e3i.cell(row=row, column=col, value=val)
        c.fill = fill
        c.font = normal
        c.alignment = Alignment(wrap_text=(col in (2, 7, 8)), vertical="center")

col_widths_e3 = [16, 50, 28, 6, 14, 12, 45, 35, 40, 30, 18, 14, 10]
for i, w in enumerate(col_widths_e3, start=1):
    ws_e3i.column_dimensions[get_column_letter(i)].width = w

ws_e3i.freeze_panes = "A2"

# ════════════════════════════════════════════════════════════════════════════
# HOJA 5: ETAPA 3 — EXCLUIDOS (relevancia < 0.4) + sin PDF
# ════════════════════════════════════════════════════════════════════════════
ws_e3x = wb.create_sheet("Etapa3_Excluidos")

headers_e3x = ["ID", "Título", "Autores", "Año",
                "Relevancia E3", "Decisión", "Motivo Exclusión", "Criterio"]

for col, h in enumerate(headers_e3x, start=1):
    c = ws_e3x.cell(row=1, column=col, value=h)
    c.fill = PatternFill("solid", fgColor="C00000")
    c.font = bold_white
    c.alignment = Alignment(horizontal="center", vertical="center")

excluidos_e3 = [a for a in arts_e3 if a.get("relevancia_general", 0) < 0.4]

all_excluded_e3 = []
for art in excluidos_e3:
    all_excluded_e3.append({**art, "_motivo": f"Relevancia < 0.4 en análisis texto completo ({art.get('relevancia_general','')})", "_criterio": "CE1"})
for art in sin_pdf:
    all_excluded_e3.append({**art, "_motivo": "Texto completo no disponible (acceso restringido o sin respuesta de autor)", "_criterio": "CE5"})

for row, art in enumerate(all_excluded_e3, start=2):
    vals = [
        art.get("id", ""),
        art.get("titulo", ""),
        get_autores(art),
        art.get("year", ""),
        art.get("relevancia_general", "N/A"),
        "EXCLUIDO",
        art["_motivo"],
        art["_criterio"],
    ]
    fill = PatternFill("solid", fgColor="FFF2CC") if row % 2 == 0 else blanco_fill
    for col, val in enumerate(vals, start=1):
        c = ws_e3x.cell(row=row, column=col, value=val)
        c.fill = fill
        c.font = normal
        c.alignment = Alignment(wrap_text=(col in (2, 7)), vertical="center")

col_widths_e3x = [16, 55, 28, 6, 14, 12, 60, 10]
for i, w in enumerate(col_widths_e3x, start=1):
    ws_e3x.column_dimensions[get_column_letter(i)].width = w

ws_e3x.freeze_panes = "A2"

# ── Guardar ───────────────────────────────────────────────────────────────────
output_path = f"{BASE}/prisma/SCREENING_DECISIONS_GENERADO.xlsx"
wb.save(output_path)
print(f"\nArchivo guardado: {output_path}")
print(f"\nResumen hojas:")
print(f"  Resumen PRISMA          : 1 tabla")
print(f"  Etapa2_Elegibles        : {len(elegibles_e2)} filas")
print(f"  Etapa2_Excluidos        : {len(excluidos_e2)} filas")
print(f"  Etapa3_Incluidos        : {len(incluidos_e3)} filas")
print(f"  Etapa3_Excluidos        : {len(all_excluded_e3)} filas  ({len(excluidos_e3)} baja relevancia + {len(sin_pdf)} sin PDF)")
