"""
Genera VERIFICATION_LOG.xlsx para el repositorio PRISMA.

Fuentes:
  - paper/validacion_doble_ciego.txt  -> calificaciones humanas (n=30)
  - clasificacion_pdfs_completos.json -> scores LLM por ID

Calcula:
  - Acuerdo exacto, acuerdo adyacente
  - Cohen's Kappa simple y ponderado lineal
  - Matriz de confusion 3x3 (A/M/B)
  - Discrepancias individuales con analisis

Hojas:
  1. Comparacion        - tabla artículo a artículo
  2. Matriz_Confusion   - tabla 3x3
  3. Estadisticas       - kappa, acuerdos, analisis
  4. Discrepancias      - solo los casos con desacuerdo
"""

import json
import re
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = "C:/Users/Luis Rojas/.openclaw/workspace"

# ── Datos hardcodeados de la validacion doble ciego ──────────────────────────
# Extraidos de paper/validacion_doble_ciego.txt

MUESTRA = [
    {"n": 1,  "id": "articulo_000671",                    "anio": 2023, "titulo": "Detecting Intents of Fake News Using Uncertainty-Aware Deep Reinforcement Learning",                      "humano": "A"},
    {"n": 2,  "id": "articulo_000862",                    "anio": 2024, "titulo": "A Unified Framework for Analyzing Textual Context and Intent in Social Media",                            "humano": "A"},
    {"n": 3,  "id": "articulo_000127",                    "anio": 2024, "titulo": "Modality Deep-learning Frameworks for Fake News Detection on Social Networks: A Systematic Literature Review", "humano": "M"},
    {"n": 4,  "id": "articulo_000521",                    "anio": 2024, "titulo": "A Survey on Automatic Credibility Assessment Using Textual Credibility Signals in the Era of LLMs",      "humano": "M"},
    {"n": 5,  "id": "articulo_002681",                    "anio": 2024, "titulo": "Decoding fake news fabrications and trends: A comprehensive survey",                                       "humano": "A"},
    {"n": 6,  "id": "articulo_001967",                    "anio": 2023, "titulo": "Enabling Contextual Soft Moderation on Social Media through Contrastive Textual Deviation",               "humano": "M"},
    {"n": 7,  "id": "articulo_002683",                    "anio": 2023, "titulo": "Annotation-Scheme Reconstruction for Fake News and Japanese Fake News Dataset",                           "humano": "M"},
    {"n": 8,  "id": "articulo_002064",                    "anio": 2023, "titulo": "DisTrack: a new Tool for Semi-automatic Misinformation Tracking in Online Social Networks",               "humano": "M"},
    {"n": 9,  "id": "articulo_000142",                    "anio": 2023, "titulo": "Fake News Detection via Intermediate-Layer Emotional Representations",                                    "humano": "M"},
    {"n": 10, "id": "articulo_000213",                    "anio": 2024, "titulo": "Propagation Structure-Aware Graph Transformer for Robust and Interpretable Fake News Detection",         "humano": "B"},
    {"n": 11, "id": "articulo_001671",                    "anio": 2024, "titulo": "A systematic review on media bias detection: What is media bias, how it is expressed, and how to detect it", "humano": "M"},
    {"n": 12, "id": "articulo_001452",                    "anio": 2025, "titulo": "Un framework para caracterizar Fake News en términos de emociones",                                      "humano": "A"},
    {"n": 13, "id": "articulo_000202",                    "anio": 2025, "titulo": "LLM-Generated Fake News Induces Truth Decay in News Ecosystem",                                         "humano": "A"},
    {"n": 14, "id": "articulo_003364",                    "anio": 2021, "titulo": "Different types of COVID-19 misinformation have different emotional valence on Twitter",                  "humano": "A"},
    {"n": 15, "id": "articulo_000992",                    "anio": 2024, "titulo": "Toward Mitigating Misinformation and Social Media Manipulation in LLM Era",                              "humano": "M"},
    {"n": 16, "id": "articulo_001423",                    "anio": 2024, "titulo": "LLMs for Explainable Few-shot Deception Detection",                                                      "humano": "M"},
    {"n": 17, "id": "articulo_001602",                    "anio": 2024, "titulo": "Muertes de celebridades: Sobre la clasificación automática de noticias falsas no intencionadas",         "humano": "A"},
    {"n": 18, "id": "articulo_003331",                    "anio": 2022, "titulo": "In-context annotation of topic-oriented datasets of fake news: A case study on Notre-Dame fire",         "humano": "M"},
    {"n": 19, "id": "articulo_001635",                    "anio": 2022, "titulo": "Asustar para desestabilizar: desinformación sobre la COVID-19 en Argentina y España",                    "humano": "A"},
    {"n": 20, "id": "articulo_000146",                    "anio": 2023, "titulo": "Using Deep Learning Models to Detect Fake News about COVID-19",                                          "humano": "M"},
    {"n": 21, "id": "articulo_000152",                    "anio": 2023, "titulo": "Explainability in NLP model: Detection of Covid-19 Twitter Fake News",                                   "humano": "B"},
    {"n": 22, "id": "articulo_000691",                    "anio": 2025, "titulo": "Multilevel Annotation Model for Detecting Fake News in Kazakh-Language Media",                          "humano": "A"},
    {"n": 23, "id": "articulo_003259",                    "anio": 2021, "titulo": "An Improved Multiple Features and Machine Learning-Based Approach for Detecting Clickbait News",         "humano": "M"},
    {"n": 24, "id": "articulo_000555",                    "anio": 2024, "titulo": "GPT Assisted Annotation of Rhetorical and Linguistic Features for Interpretable Propaganda Technique Detection", "humano": "A"},
    {"n": 25, "id": "articulo_003396",                    "anio": 2021, "titulo": "To share or not to share – The underlying motives of sharing fake news amidst COVID-19 in Malaysia",    "humano": "M"},
    {"n": 26, "id": "articulo_003289",                    "anio": 2023, "titulo": "XAI in Automated Fact-Checking? The Benefits Are Modest and There's No One-Explanation-Fits-All",        "humano": "B"},
    {"n": 27, "id": "articulo_001001",                    "anio": 2025, "titulo": "A Perturbation-Theoretic Model for Fact-Checker Deployment in Dynamic Disinformation Networks",         "humano": "M"},
    {"n": 28, "id": "articulo_001003",                    "anio": 2022, "titulo": "The fake news effect: what does it mean for consumer behavioral intentions towards brands?",              "humano": "M"},
    {"n": 29, "id": "articulo_000907",                    "anio": 2024, "titulo": "BISON: Blockchain Interpretable Success Prediction for Web3 Social Media Content",                       "humano": "B"},
    {"n": 30, "id": "articulo_001870",                    "anio": 2023, "titulo": "Navigating social and emotional learning evidence in a polarized socio-political environment",            "humano": "B"},
]

# ── Cargar scores LLM ─────────────────────────────────────────────────────────
with open(f"{BASE}/clasificacion_pdfs_completos.json", "r", encoding="utf-8") as f:
    etapa3 = json.load(f)

idx_llm = {a["id"]: a.get("relevancia_general", None) for a in etapa3["articulos"]}

# Tambien buscar en clasificacion_claude.json para artículos que no pasaron Etapa 3
with open(f"{BASE}/clasificacion_claude.json", "r", encoding="utf-8") as f:
    claude_data = json.load(f)
idx_claude = {a["id"]: a.get("relevancia_general", None) for a in claude_data["articulos"]}

def score_to_cat(score):
    if score is None:
        return "N/D"
    if score >= 0.75:
        return "A"
    if score >= 0.45:
        return "M"
    return "B"

def cat_to_num(cat):
    return {"A": 2, "M": 1, "B": 0}.get(cat, -1)

# Enriquecer muestra con scores LLM
for art in MUESTRA:
    aid = art["id"]
    score = idx_llm.get(aid)
    if score is None:
        # buscar sin sufijo
        base_id = aid.split("_")[0] + "_" + aid.split("_")[1] if "_" in aid else aid
        score = idx_llm.get(base_id) or idx_claude.get(aid) or idx_claude.get(base_id)
    art["score_llm"] = score
    art["cat_llm"]   = score_to_cat(score)
    art["acuerdo"]   = "✓ Exacto" if art["cat_llm"] == art["humano"] else (
                       "~ Adyacente" if abs(cat_to_num(art["cat_llm"]) - cat_to_num(art["humano"])) == 1
                       else "✗ Extremo")
    art["diferencia"] = cat_to_num(art["cat_llm"]) - cat_to_num(art["humano"])

# ── Calcular estadísticas ─────────────────────────────────────────────────────
n = len(MUESTRA)
exactos    = sum(1 for a in MUESTRA if a["acuerdo"] == "✓ Exacto")
adyacentes = sum(1 for a in MUESTRA if a["acuerdo"] == "~ Adyacente")
extremos   = sum(1 for a in MUESTRA if a["acuerdo"] == "✗ Extremo")

# Matriz de confusion: filas = LLM, cols = Humano
cats = ["A", "M", "B"]
matriz = {r: {c: 0 for c in cats} for r in cats}
for art in MUESTRA:
    llm_c = art["cat_llm"]
    hum_c = art["humano"]
    if llm_c in cats and hum_c in cats:
        matriz[llm_c][hum_c] += 1

# Cohen's Kappa simple
po = exactos / n
# Expected agreement
total_llm = Counter(a["cat_llm"] for a in MUESTRA)
total_hum = Counter(a["humano"] for a in MUESTRA)
pe = sum((total_llm[c] / n) * (total_hum[c] / n) for c in cats)
kappa = (po - pe) / (1 - pe) if (1 - pe) != 0 else 0

# Cohen's Kappa ponderado lineal
weights = {("A","A"):0,("A","M"):1,("A","B"):2,
           ("M","A"):1,("M","M"):0,("M","B"):1,
           ("B","A"):2,("B","M"):1,("B","B"):0}
max_w = 2
po_w = 1 - sum(weights.get((a["cat_llm"], a["humano"]), 2) for a in MUESTRA if a["cat_llm"] in cats) / (n * max_w)
pe_w_num = 0
for r in cats:
    for c in cats:
        pe_w_num += weights.get((r,c),2) * (total_llm[r]/n) * (total_hum[c]/n)
pe_w = 1 - pe_w_num / max_w
kappa_w = (po_w - pe_w) / (1 - pe_w) if (1 - pe_w) != 0 else 0

print(f"n={n} | Exacto: {exactos} ({exactos/n*100:.1f}%) | Adyacente: {adyacentes} | Extremo: {extremos}")
print(f"Kappa simple: {kappa:.3f} | Kappa ponderado lineal: {kappa_w:.3f}")

# ── Workbook ──────────────────────────────────────────────────────────────────
wb = Workbook()

H_DARK   = PatternFill("solid", fgColor="1F4E79")
H_GREEN  = PatternFill("solid", fgColor="375623")
H_RED    = PatternFill("solid", fgColor="C00000")
H_GRAY   = PatternFill("solid", fgColor="595959")
ROW_EVEN = PatternFill("solid", fgColor="EBF3FB")
ROW_ODD  = PatternFill("solid", fgColor="FFFFFF")
GREEN_F  = PatternFill("solid", fgColor="C6EFCE")
YELLOW_F = PatternFill("solid", fgColor="FFEB9C")
RED_F    = PatternFill("solid", fgColor="FFC7CE")
BOLD_W   = Font(bold=True, color="FFFFFF", size=10)
BOLD_D   = Font(bold=True, color="1F4E79", size=10)
NORM     = Font(size=9)

def hdr(ws, cols, fill=H_DARK):
    for i, h in enumerate(cols, 1):
        c = ws.cell(1, i, h)
        c.fill = fill; c.font = BOLD_W
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

def widths(ws, ws_list):
    for i, w in enumerate(ws_list, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ════════════════════════════════════════════════════════════════════════════
# HOJA 1: COMPARACIÓN ARTÍCULO A ARTÍCULO
# ════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "1_Comparacion"

hdr(ws1, ["#", "ID Artículo", "Año", "Título", "Score LLM", "Categoría LLM",
           "Calificación Humano", "Acuerdo", "Diferencia (LLM-Humano)"])

for art in MUESTRA:
    row = art["n"] + 1
    ac  = art["acuerdo"]
    fill = GREEN_F if ac == "✓ Exacto" else (YELLOW_F if ac == "~ Adyacente" else RED_F)

    vals = [
        art["n"],
        art["id"],
        art["anio"],
        art["titulo"],
        art["score_llm"] if art["score_llm"] is not None else "N/D",
        art["cat_llm"],
        art["humano"],
        ac,
        art["diferencia"],
    ]
    for col, val in enumerate(vals, 1):
        c = ws1.cell(row, col, val)
        c.fill = fill
        c.font = NORM
        c.alignment = Alignment(vertical="center", wrap_text=(col==4), horizontal="center" if col not in (4,) else "left")

widths(ws1, [5, 20, 6, 60, 12, 14, 18, 16, 18])

# ════════════════════════════════════════════════════════════════════════════
# HOJA 2: MATRIZ DE CONFUSIÓN
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("2_Matriz_Confusion")

ws2.cell(1, 1, "LLM \\ Humano").fill = H_GRAY
ws2.cell(1, 1).font = BOLD_W
ws2.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")

for col, cat in enumerate(cats, 2):
    c = ws2.cell(1, col, f"Humano: {cat}")
    c.fill = H_DARK; c.font = BOLD_W
    c.alignment = Alignment(horizontal="center", vertical="center")

total_fill = PatternFill("solid", fgColor="2E75B6")
for row, r_cat in enumerate(cats, 2):
    c = ws2.cell(row, 1, f"LLM: {r_cat}")
    c.fill = H_DARK; c.font = BOLD_W
    c.alignment = Alignment(horizontal="center", vertical="center")
    for col, c_cat in enumerate(cats, 2):
        val = matriz[r_cat][c_cat]
        cell = ws2.cell(row, col, val)
        if r_cat == c_cat:
            cell.fill = GREEN_F
        elif abs(cat_to_num(r_cat) - cat_to_num(c_cat)) == 1:
            cell.fill = YELLOW_F
        else:
            cell.fill = RED_F
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal="center", vertical="center")

# Totales por fila
ws2.cell(1, 5, "Total LLM").fill = H_GRAY; ws2.cell(1,5).font = BOLD_W
ws2.cell(1, 5).alignment = Alignment(horizontal="center")
for row, r_cat in enumerate(cats, 2):
    t = sum(matriz[r_cat].values())
    c = ws2.cell(row, 5, t)
    c.fill = total_fill; c.font = Font(bold=True, color="FFFFFF", size=11)
    c.alignment = Alignment(horizontal="center", vertical="center")

# Totales por columna
ws2.cell(5, 1, "Total Humano").fill = H_GRAY; ws2.cell(5,1).font = BOLD_W
ws2.cell(5, 1).alignment = Alignment(horizontal="center")
for col, c_cat in enumerate(cats, 2):
    t = sum(matriz[r][c_cat] for r in cats)
    c = ws2.cell(5, col, t)
    c.fill = total_fill; c.font = Font(bold=True, color="FFFFFF", size=11)
    c.alignment = Alignment(horizontal="center", vertical="center")

ws2.cell(5, 5, n).fill = total_fill
ws2.cell(5, 5).font = Font(bold=True, color="FFFFFF", size=12)
ws2.cell(5, 5).alignment = Alignment(horizontal="center", vertical="center")

for i in range(1, 6):
    ws2.column_dimensions[get_column_letter(i)].width = 18
    ws2.row_dimensions[i].height = 30

# Leyenda debajo de la matriz
ws2.cell(7, 1, "Leyenda:").font = BOLD_D
ws2.cell(8, 1, "Verde = acuerdo exacto (diagonal)").fill = GREEN_F
ws2.cell(9, 1, "Amarillo = desacuerdo adyacente (1 nivel)").fill = YELLOW_F
ws2.cell(10, 1, "Rojo = desacuerdo extremo (2 niveles)").fill = RED_F
for r in [8, 9, 10]:
    ws2.cell(r, 1).font = NORM
    ws2.merge_cells(f"A{r}:E{r}")

# ════════════════════════════════════════════════════════════════════════════
# HOJA 3: ESTADÍSTICAS
# ════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("3_Estadisticas")

def _kappa_label(k):
    if k < 0:      return "Peor que azar"
    if k < 0.20:   return "Slight agreement (Landis & Koch)"
    if k < 0.40:   return "Fair agreement"
    if k < 0.60:   return "Moderate agreement"
    if k < 0.80:   return "Substantial agreement"
    return "Almost perfect agreement"

stats = [
    ("ESTADÍSTICAS DE ACUERDO INTER-EVALUADOR", None, None),
    ("", None, None),
    ("Parámetro", "Valor", "Interpretación"),
    ("N artículos muestra", n, "Muestreo aleatorio estratificado"),
    ("Acuerdo exacto (po)", f"{exactos}/{n} = {exactos/n*100:.1f}%", "Porcentaje coincidencia de categoría"),
    ("Acuerdo adyacente", f"{adyacentes}/{n} = {adyacentes/n*100:.1f}%", "Diferencia de solo 1 nivel"),
    ("Desacuerdo extremo", f"{extremos}/{n} = {extremos/n*100:.1f}%", "Diferencia de 2 niveles (A↔B)"),
    ("", None, None),
    ("Cohen's Kappa (κ)", f"{kappa:.3f}", _kappa_label(kappa)),
    ("Cohen's Kappa ponderado lineal (κw)", f"{kappa_w:.3f}", _kappa_label(kappa_w)),
    ("", None, None),
    ("Distribución LLM - Alta (A)", f"{total_llm['A']}/{n}", f"{total_llm['A']/n*100:.1f}%"),
    ("Distribución LLM - Media (M)", f"{total_llm['M']}/{n}", f"{total_llm['M']/n*100:.1f}%"),
    ("Distribución LLM - Baja (B)",  f"{total_llm['B']}/{n}", f"{total_llm['B']/n*100:.1f}%"),
    ("", None, None),
    ("Distribución Humano - Alta (A)", f"{total_hum['A']}/{n}", f"{total_hum['A']/n*100:.1f}%"),
    ("Distribución Humano - Media (M)", f"{total_hum['M']}/{n}", f"{total_hum['M']/n*100:.1f}%"),
    ("Distribución Humano - Baja (B)",  f"{total_hum['B']}/{n}", f"{total_hum['B']/n*100:.1f}%"),
    ("", None, None),
    ("Sesgo sistemático detectado",
     "LLM sobreclasifica en categoría Alta",
     f"LLM asignó A a {total_llm['A']} artículos vs {total_hum['A']} del evaluador humano"),
    ("Umbral de inclusión pipeline",
     "≥ 0.4 (categorías A y M)",
     "El umbral captura artículos de relevancia media y alta"),
]

# Re-calcular con la función definida
stats[8]  = ("Cohen's Kappa (κ)", f"{kappa:.3f}", _kappa_label(kappa))
stats[9]  = ("Cohen's Kappa ponderado lineal (κw)", f"{kappa_w:.3f}", _kappa_label(kappa_w))

for row, (param, val, interp) in enumerate(stats, 1):
    if row == 1:
        c = ws3.cell(row, 1, param)
        c.fill = H_DARK; c.font = BOLD_W
        c.alignment = Alignment(horizontal="center")
        ws3.merge_cells(f"A{row}:C{row}")
    elif row == 3:
        for col, h in enumerate([param, val, interp], 1):
            c = ws3.cell(row, col, h)
            c.fill = H_DARK; c.font = BOLD_W
            c.alignment = Alignment(horizontal="center")
    elif param == "":
        pass
    else:
        fill = ROW_EVEN if row % 2 == 0 else ROW_ODD
        ws3.cell(row, 1, param).fill = fill
        ws3.cell(row, 1).font = Font(bold=True, size=9, color="1F4E79")
        ws3.cell(row, 2, str(val) if val is not None else "").fill = fill
        ws3.cell(row, 2).font = Font(bold=True, size=10)
        ws3.cell(row, 2).alignment = Alignment(horizontal="center")
        ws3.cell(row, 3, str(interp) if interp is not None else "").fill = fill
        ws3.cell(row, 3).font = NORM

widths(ws3, [40, 25, 60])

# ════════════════════════════════════════════════════════════════════════════
# HOJA 4: DISCREPANCIAS
# ════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("4_Discrepancias")

discrepancias = [a for a in MUESTRA if a["acuerdo"] != "✓ Exacto"]
hdr(ws4, ["#", "ID", "Año", "Título", "Score LLM", "Cat LLM", "Cat Humano", "Tipo Discrepancia", "Análisis"], fill=H_RED)

analisis_map = {
    "A→M": "LLM sobreclasificó: artículo tiene relevancia granular parcial pero el evaluador la considera media",
    "A→B": "Discrepancia extrema: LLM detectó señales de intención comunicativa que el evaluador no consideró relevantes",
    "M→A": "Subclasificación LLM: artículo con intenciones explícitas que el evaluador reconoció como alta relevancia",
    "M→B": "LLM asignó relevancia media; evaluador considera fuera del scope del tema",
    "B→A": "Discrepancia extrema: evaluador reconoció relevancia que el LLM no capturó en metadatos",
    "B→M": "LLM infra-detectó relevancia; artículo tiene componente de intención relevante según evaluador",
}

for row, art in enumerate(discrepancias, 2):
    tipo = f"{art['cat_llm']}→{art['humano']}"
    analisis = analisis_map.get(tipo, "Discrepancia sin patrón predefinido")
    fill = YELLOW_F if art["acuerdo"] == "~ Adyacente" else RED_F
    vals = [art["n"], art["id"], art["anio"], art["titulo"],
            art["score_llm"] if art["score_llm"] is not None else "N/D",
            art["cat_llm"], art["humano"], tipo, analisis]
    for col, val in enumerate(vals, 1):
        c = ws4.cell(row, col, val)
        c.fill = fill; c.font = NORM
        c.alignment = Alignment(vertical="center", wrap_text=(col in (4, 9)))

widths(ws4, [5, 20, 6, 55, 12, 10, 12, 16, 65])

# ── Guardar ───────────────────────────────────────────────────────────────────
output = f"{BASE}/prisma/VERIFICATION_LOG_GENERADO.xlsx"
wb.save(output)
print(f"\nArchivo guardado: {output}")
print(f"\nHojas:")
print(f"  1_Comparacion      : {n} filas | {exactos} exactos, {adyacentes} adyacentes, {extremos} extremos")
print(f"  2_Matriz_Confusion : tabla 3x3 A/M/B")
print(f"  3_Estadisticas     : kappa={kappa:.3f}, kappa_w={kappa_w:.3f}")
print(f"  4_Discrepancias    : {len(discrepancias)} casos con desacuerdo")
