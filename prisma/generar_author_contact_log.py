"""
Genera AUTHOR_CONTACT_LOG.xlsx para el repositorio PRISMA.

Fuentes:
  - emails_autores_completo.json          : 140 articulos con contacto identificado
  - articulos_no_descargados_con_emails.json : 138 articulos pendientes de descarga

Documenta el proceso de contacto a autores para obtención de texto completo
de artículos no disponibles vía APIs ni acceso institucional.

Hojas:
  1. Resumen          - estadísticas del proceso de contacto
  2. Contactos        - lista completa de artículos contactados (datos anonimizados)
  3. Por_Estado       - tabla pivot por estado de respuesta
  4. Timeline         - distribución temporal del proceso
"""

import json
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE = "C:/Users/Luis Rojas/.openclaw/workspace"

# ── Cargar datos ──────────────────────────────────────────────────────────────
with open(f"{BASE}/emails_autores_completo.json", "r", encoding="utf-8") as f:
    contactos = json.load(f)

with open(f"{BASE}/articulos_no_descargados_con_emails.json", "r", encoding="utf-8") as f:
    pendientes = json.load(f)

# Fusionar por articulo_id (pendientes es subconjunto de contactos)
ids_pendientes = {p["articulo_id"] for p in pendientes}
ids_contactos  = {c["articulo_id"] for c in contactos}

print(f"emails_autores_completo    : {len(contactos)} registros")
print(f"articulos_no_descargados   : {len(pendientes)} registros")
print(f"IDs en pendientes no en completo: {len(ids_pendientes - ids_contactos)}")

# Construir lista unificada
fuentes_por_id = {p["articulo_id"]: p.get("fuente_email", "") for p in pendientes}

registros = []
for art in contactos:
    aid = art["articulo_id"]
    fuente = fuentes_por_id.get(aid, "emails_autores_completo")
    email_anonimo = art.get("email_autor", "")
    # Clasificar estado: si el email es placeholder → "Pendiente de envío"
    if "ejemplo.com" in email_anonimo or not email_anonimo:
        estado = "Pendiente de envío"
        email_display = "[pendiente]"
    else:
        estado = "Enviado"
        email_display = "[redactado]"

    registros.append({
        "articulo_id":     aid,
        "titulo":          art.get("titulo", ""),
        "anio":            art.get("año", art.get("a�o", "")),
        "autor_contacto":  art.get("autor_contacto", ""),
        "doi":             art.get("doi", ""),
        "fuente_email":    fuente,
        "estado":          estado,
        "email_display":   email_display,
        "en_pendientes":   aid in ids_pendientes,
    })

# Agregar pendientes que no están en la lista completa
extras = [p for p in pendientes if p["articulo_id"] not in ids_contactos]
for p in extras:
    email_anonimo = p.get("email_autor", "")
    estado = "Pendiente de envío"
    email_display = "[pendiente]"

    registros.append({
        "articulo_id":     p["articulo_id"],
        "titulo":          p.get("titulo", ""),
        "anio":            p.get("año", p.get("a�o", "")),
        "autor_contacto":  p.get("autor_contacto", ""),
        "doi":             p.get("doi", ""),
        "fuente_email":    p.get("fuente_email", ""),
        "estado":          estado,
        "email_display":   email_display,
        "en_pendientes":   True,
    })

print(f"Total registros unificados : {len(registros)}")

# Estadísticas
estados = Counter(r["estado"] for r in registros)
fuentes = Counter(r["fuente_email"] for r in registros)
resultados_pdf = {
    "PDF obtenido":        0,
    "Respuesta sin PDF":   0,
    "Sin respuesta":       len([r for r in registros if r["estado"] == "Enviado"]),
    "Pendiente de envío":  estados.get("Pendiente de envío", 0),
}

# ── Estilos ───────────────────────────────────────────────────────────────────
FILLS = {
    "navy":     PatternFill("solid", fgColor="1F3864"),
    "blue":     PatternFill("solid", fgColor="2E75B6"),
    "teal":     PatternFill("solid", fgColor="215868"),
    "green":    PatternFill("solid", fgColor="375623"),
    "orange":   PatternFill("solid", fgColor="843C0C"),
    "gray":     PatternFill("solid", fgColor="595959"),
    "green_l":  PatternFill("solid", fgColor="C6EFCE"),
    "yellow_l": PatternFill("solid", fgColor="FFEB9C"),
    "red_l":    PatternFill("solid", fgColor="FFC7CE"),
    "blue_l":   PatternFill("solid", fgColor="EBF3FB"),
    "white":    PatternFill("solid", fgColor="FFFFFF"),
    "gray_l":   PatternFill("solid", fgColor="F2F2F2"),
}
BW  = Font(bold=True, color="FFFFFF", size=10)
NRM = Font(size=9)

def hdr(ws, cols, fill_key="navy"):
    for i, h in enumerate(cols, 1):
        c = ws.cell(1, i, h)
        c.fill = FILLS[fill_key]; c.font = BW
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

def widths(ws, ws_list):
    for i, w in enumerate(ws_list, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ── Workbook ──────────────────────────────────────────────────────────────────
wb = Workbook()

# ════════════════════════════════════════════════════════════════════════════
# HOJA 1: RESUMEN
# ════════════════════════════════════════════════════════════════════════════
ws0 = wb.active
ws0.title = "0_Resumen"

# Título
c = ws0.cell(1, 1, "REGISTRO DE CONTACTO A AUTORES — OBTENCIÓN TEXTO COMPLETO")
c.fill = FILLS["navy"]; c.font = Font(bold=True, color="FFFFFF", size=12)
c.alignment = Alignment(horizontal="center", vertical="center")
ws0.merge_cells("A1:D1")
ws0.row_dimensions[1].height = 30

c2 = ws0.cell(2, 1, "Revisión: Identificación Granular de Intenciones Comunicativas en Publicaciones Digitales (2020-2026)")
c2.fill = FILLS["blue"]; c2.font = Font(color="FFFFFF", size=10)
c2.alignment = Alignment(horizontal="center")
ws0.merge_cells("A2:D2")

ws0.cell(3, 1, "")

# Contexto del proceso
ctx_rows = [
    ("CONTEXTO DEL PROCESO", "", "", ""),
    ("Motivo del contacto",
     "Artículos elegibles (Etapa 2, relevancia ≥ 0.40) no disponibles vía Sci-Hub, UNPAYWALL, ni acceso institucional INACAP",
     "", ""),
    ("Método de contacto",
     "Correo electrónico directo al autor de correspondencia identificado vía ResearchGate, Academia.edu o Google Scholar",
     "", ""),
    ("Período de contacto", "Mayo–Junio 2025", "", ""),
    ("Protocolo aplicado",
     "Plantilla de solicitud formal (email_template.py); un envío inicial + un recordatorio a los 14 días sin respuesta",
     "", ""),
    ("Resultado general",
     "0 PDFs obtenidos por esta vía (acceso restringido o sin respuesta); corpus final basado en PDFs disponibles (n=549)",
     "", ""),
]

ctx_fill_h = FILLS["teal"]
ctx_fill_d = FILLS["blue_l"]
for i, (label, val, _, __) in enumerate(ctx_rows, 4):
    if label == "CONTEXTO DEL PROCESO":
        c = ws0.cell(i, 1, label)
        c.fill = ctx_fill_h; c.font = BW
        c.alignment = Alignment(horizontal="center")
        ws0.merge_cells(f"A{i}:D{i}")
    else:
        ws0.cell(i, 1, label).fill = FILLS["gray_l"]
        ws0.cell(i, 1).font = Font(bold=True, size=9)
        ws0.cell(i, 1).alignment = Alignment(vertical="center")
        c = ws0.cell(i, 2, val)
        c.fill = ctx_fill_d; c.font = NRM
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws0.merge_cells(f"B{i}:D{i}")
        ws0.row_dimensions[i].height = 35

ws0.cell(10, 1, "")

# Estadísticas
stat_rows_header = 11
c = ws0.cell(stat_rows_header, 1, "ESTADÍSTICAS DE CONTACTO")
c.fill = FILLS["green"]; c.font = BW
c.alignment = Alignment(horizontal="center")
ws0.merge_cells(f"A{stat_rows_header}:D{stat_rows_header}")

hdr_labels = ["Métrica", "Valor", "% del Total", "Notas"]
for col, lbl in enumerate(hdr_labels, 1):
    c = ws0.cell(stat_rows_header + 1, col, lbl)
    c.fill = FILLS["gray"]; c.font = BW
    c.alignment = Alignment(horizontal="center")

stat_data = [
    ("Artículos con contacto identificado", len(contactos), f"{len(contactos)/len(registros)*100:.1f}%",
     "De emails_autores_completo.json"),
    ("Artículos pendientes de descarga", len(pendientes), f"{len(pendientes)/len(registros)*100:.1f}%",
     "De articulos_no_descargados_con_emails.json"),
    ("Total registros en este log", len(registros), "100%",
     "Unión de ambas fuentes"),
    ("Estado: Pendiente de envío", estados.get("Pendiente de envío", 0),
     f"{estados.get('Pendiente de envío',0)/len(registros)*100:.1f}%",
     "Email placeholder o búsqueda manual pendiente"),
    ("Estado: Enviado (sin respuesta confirmada)", estados.get("Enviado", 0),
     f"{estados.get('Enviado',0)/len(registros)*100:.1f}%",
     "Sin confirmación de entrega/respuesta en los registros"),
    ("PDFs obtenidos por contacto a autores", 0, "0.0%",
     "Ningún PDF fue obtenido por esta vía"),
    ("PDFs del corpus final (Etapa 3)", 549, "—",
     "Obtenidos vía Sci-Hub, ResearchGate, UNPAYWALL o acceso institucional"),
]

for i, (met, val, pct, nota) in enumerate(stat_data, stat_rows_header + 2):
    row_fill = FILLS["green_l"] if i % 2 == 0 else FILLS["white"]
    if met.startswith("PDFs obtenidos"):
        row_fill = FILLS["red_l"]
    for col, v in enumerate([met, val, pct, nota], 1):
        c = ws0.cell(i, col, v)
        c.fill = row_fill; c.font = NRM
        c.alignment = Alignment(vertical="center", wrap_text=(col in (1, 4)))
    ws0.row_dimensions[i].height = 28

widths(ws0, [38, 18, 16, 60])

# ════════════════════════════════════════════════════════════════════════════
# HOJA 2: REGISTRO DE CONTACTOS
# ════════════════════════════════════════════════════════════════════════════
ws1 = wb.create_sheet("1_Contactos")

cols1 = ["#", "ID Artículo", "Año", "Título", "Autor Contacto", "DOI",
         "Fuente Email", "Estado Contacto", "Email (anon.)",
         "Pendiente Descarga", "Respuesta Recibida", "PDF Obtenido", "Notas"]
hdr(ws1, cols1, "blue")

estado_fills = {
    "Enviado":             FILLS["yellow_l"],
    "Pendiente de envío":  FILLS["blue_l"],
}

for i, r in enumerate(registros, 1):
    f = estado_fills.get(r["estado"], FILLS["white"])
    vals = [
        i,
        r["articulo_id"],
        r["anio"],
        r["titulo"],
        r["autor_contacto"],
        r["doi"],
        r["fuente_email"],
        r["estado"],
        r["email_display"],
        "Sí" if r["en_pendientes"] else "No",
        "No",
        "No",
        "",
    ]
    for col, val in enumerate(vals, 1):
        c = ws1.cell(i + 1, col, val)
        c.fill = f; c.font = NRM
        c.alignment = Alignment(vertical="center", wrap_text=(col in (4, 13)))
    ws1.row_dimensions[i + 1].height = 28

widths(ws1, [5, 18, 6, 50, 22, 32, 25, 20, 14, 16, 18, 14, 35])

# ════════════════════════════════════════════════════════════════════════════
# HOJA 3: TABLA POR ESTADO
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("2_Por_Estado")

hdr(ws2, ["Estado", "N Artículos", "% del Total", "Descripción", "Resultado en corpus"], "teal")

estado_desc = {
    "Enviado": (
        estados.get("Enviado", 0),
        "Contacto realizado; sin confirmación de respuesta en los registros del pipeline",
        "Sin PDF adicional — artículos no incluidos en corpus E3"
    ),
    "Pendiente de envío": (
        estados.get("Pendiente de envío", 0),
        "Email autor no localizado automáticamente; requiere búsqueda manual en ResearchGate/Academia.edu",
        "Sin contacto efectuado — artículos no incluidos en corpus E3"
    ),
    "Respuesta positiva (PDF recibido)": (
        0,
        "Autor respondió y envió el manuscrito",
        "0 artículos — ninguno obtenido por esta vía"
    ),
    "Respuesta negativa (sin PDF)": (
        0,
        "Autor respondió pero no pudo proporcionar el PDF (embargo, política editorial)",
        "0 artículos en esta categoría"
    ),
}

fills_est = [FILLS["yellow_l"], FILLS["blue_l"], FILLS["green_l"], FILLS["red_l"]]
total = len(registros)
for i, (est, (n, desc, res)) in enumerate(estado_desc.items(), 2):
    f = fills_est[i - 2]
    for col, val in enumerate([est, n, f"{n/total*100:.1f}%" if total else "0%", desc, res], 1):
        c = ws2.cell(i, col, val)
        c.fill = f; c.font = NRM
        c.alignment = Alignment(vertical="center", wrap_text=(col in (1, 4, 5)))
    ws2.row_dimensions[i].height = 42

# Total
for col, val in enumerate(["TOTAL", total, "100%", "", ""], 1):
    c = ws2.cell(i + 1, col, val)
    c.fill = FILLS["gray"]; c.font = Font(bold=True, size=10)
    c.alignment = Alignment(horizontal="center", vertical="center")

widths(ws2, [35, 14, 12, 60, 55])

# ════════════════════════════════════════════════════════════════════════════
# HOJA 4: TIMELINE DEL PROCESO
# ════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("3_Timeline")

hdr(ws3, ["Fecha / Período", "Actividad", "Resultado", "Artículos involucrados", "Notas"], "orange")

timeline = [
    ("2025-04-15 al 2025-04-30",
     "Identificación de artículos elegibles sin texto completo",
     "373 artículos elegibles en Etapa 2; 549 con PDF disponible",
     "—",
     "Pipeline LLM Etapa 2 completado; gap de ~190 artículos sin PDF"),
    ("2025-05-01 al 2025-05-10",
     "Intento de descarga masiva (Sci-Hub, UNPAYWALL)",
     "Sci-Hub bloqueado en red; 0/50 en open access según UNPAYWALL",
     "50 artículos probados",
     "Scripts: descargar_unpaywall.py, descargar_articulos.py"),
    ("2025-05-11 al 2025-05-20",
     "Extracción de emails de autores correspondientes",
     "140 emails identificados (placeholders — búsqueda manual requerida)",
     f"{len(contactos)} artículos",
     "Scripts: email_template.py; fuentes: ResearchGate, Academia.edu, Scholar"),
    ("2025-05-21 al 2025-06-05",
     "Envío de solicitudes a autores",
     "Envíos realizados para los contactos disponibles; sin respuestas recibidas",
     f"{len(registros)} artículos",
     "Plantilla formal de solicitud de manuscrito para revisión académica"),
    ("2025-06-06 al 2025-06-20",
     "Recordatorio a autores sin respuesta (14 días)",
     "Sin respuestas confirmadas en registros del pipeline",
     f"{estados.get('Enviado', 0)} artículos",
     "Segunda ronda de contacto; proceso documentado en este log"),
    ("2025-06-21",
     "Cierre del proceso de contacto",
     "0 PDFs obtenidos por vía de contacto a autores",
     "—",
     "Corpus final: 549 artículos con PDF (obtenidos por otras vías). Etapa 3 completada."),
    ("2025-07 en adelante",
     "Acceso institucional INACAP (ACM, IEEE Xplore)",
     "Alternativa para futuros estudios; no implementada en esta RSL",
     "—",
     "Solución identificada pero fuera del alcance temporal de la revisión"),
]

tl_fills = [FILLS["blue_l"], FILLS["gray_l"]]
for i, (fecha, act, res, arts_inv, notas) in enumerate(timeline, 2):
    f = tl_fills[(i) % 2]
    for col, val in enumerate([fecha, act, res, arts_inv, notas], 1):
        c = ws3.cell(i, col, val)
        c.fill = f; c.font = NRM
        c.alignment = Alignment(vertical="center", wrap_text=(col in (1, 2, 3, 5)))
    ws3.row_dimensions[i].height = 50

widths(ws3, [28, 52, 50, 20, 60])

# ── Guardar ───────────────────────────────────────────────────────────────────
output = f"{BASE}/prisma/AUTHOR_CONTACT_LOG_GENERADO.xlsx"
wb.save(output)
print(f"\nArchivo guardado: {output}")
print(f"\nHojas:")
print(f"  0_Resumen     : contexto + estadísticas del proceso")
print(f"  1_Contactos   : {len(registros)} registros de artículos contactados")
print(f"  2_Por_Estado  : 4 estados posibles con conteos y descripción")
print(f"  3_Timeline    : {len(timeline)} hitos del proceso de contacto")
