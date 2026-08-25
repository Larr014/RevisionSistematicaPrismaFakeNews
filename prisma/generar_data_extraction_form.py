"""
Genera DATA_EXTRACTION_FORM.xlsx para el repositorio PRISMA.

Fuente: clasificacion_pdfs_completos.json (Etapa 3, n=549)
Scope:  Solo los 324 articulos con relevancia >= 0.4 (corpus de sintesis)

Hojas:
  1. Indice           - Lista de 324 articulos con datos bibliograficos
  2. Metodos          - Metodos especificos y generales por articulo
  3. Intenciones      - Dimensiones intencionales por articulo
  4. Datasets         - Datasets utilizados con referencias
  5. Metricas         - Metricas de evaluacion por articulo
  6. Caracteristicas  - Caracteristicas linguisticas y metodologicas
  7. Hallazgos        - Hallazgos emergentes (texto libre por articulo)
"""

import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE = "C:/Users/Luis Rojas/.openclaw/workspace"

# ── Cargar y filtrar ──────────────────────────────────────────────────────────
with open(f"{BASE}/clasificacion_pdfs_completos.json", "r", encoding="utf-8") as f:
    data = json.load(f)

arts = [a for a in data["articulos"] if a.get("relevancia_general", 0) >= 0.4]
print(f"Corpus: {len(arts)} articulos incluidos (relevancia >= 0.4)")

# ── Helpers ───────────────────────────────────────────────────────────────────
def safe_list(art, *keys):
    obj = art.get("clasificacion", {})
    for k in keys:
        if not isinstance(obj, dict):
            return []
        obj = obj.get(k, {})
    return obj if isinstance(obj, list) else []

def safe_str(art, *keys):
    return "; ".join(safe_list(art, *keys))

def get_autores(art):
    auths = art.get("autores", [])
    if isinstance(auths, list):
        return "; ".join(auths[:3]) + (" et al." if len(auths) > 3 else "")
    return str(auths)

def get_granularidad(intenciones):
    n = len(intenciones)
    if n == 0:
        return "Sin clasificar"
    if n == 1:
        return "Binario (1 dimensión)"
    return f"Granular ({n} dimensiones)"

# ── Estilos ───────────────────────────────────────────────────────────────────
FILLS = {
    "header_dark":  PatternFill("solid", fgColor="1F4E79"),
    "header_green": PatternFill("solid", fgColor="375623"),
    "header_teal":  PatternFill("solid", fgColor="215868"),
    "header_purple":PatternFill("solid", fgColor="4B2D83"),
    "header_brown": PatternFill("solid", fgColor="833200"),
    "header_gray":  PatternFill("solid", fgColor="595959"),
    "header_navy":  PatternFill("solid", fgColor="002060"),
    "row_even":     PatternFill("solid", fgColor="EBF3FB"),
    "row_odd":      PatternFill("solid", fgColor="FFFFFF"),
    "row_green_e":  PatternFill("solid", fgColor="E2EFDA"),
    "row_green_o":  PatternFill("solid", fgColor="FFFFFF"),
    "row_teal_e":   PatternFill("solid", fgColor="E2F0EF"),
    "row_purple_e": PatternFill("solid", fgColor="EDE7F6"),
    "row_brown_e":  PatternFill("solid", fgColor="FBE9E7"),
    "row_gray_e":   PatternFill("solid", fgColor="F2F2F2"),
}
BOLD_WHITE = Font(bold=True, color="FFFFFF", size=10)
NORMAL     = Font(size=9)

def write_headers(ws, headers, fill_key="header_dark"):
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = FILLS[fill_key]
        c.font = BOLD_WHITE
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

def set_row(ws, row, vals, even_fill, odd_fill):
    fill = even_fill if row % 2 == 0 else odd_fill
    for col, val in enumerate(vals, start=1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fill
        c.font = NORMAL
        c.alignment = Alignment(vertical="center", wrap_text=True)

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ── Workbook ──────────────────────────────────────────────────────────────────
wb = Workbook()

# ════════════════════════════════════════════════════════════════════════════
# HOJA 1: ÍNDICE BIBLIOGRÁFICO
# ════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "1_Indice"

h1 = ["#", "ID Artículo", "Título", "Autores", "Año", "Revista/Venue",
       "Relevancia E3", "Granularidad", "N Intenciones", "N Hallazgos Emergentes"]
write_headers(ws1, h1, "header_dark")

for i, art in enumerate(arts, start=1):
    intenciones = safe_list(art, "variables_principales", "intenciones")
    hallazgos   = art.get("hallazgos_nuevos", [])
    n_hall      = len(hallazgos) if isinstance(hallazgos, list) else 0
    venue       = art.get("clasificacion", {}).get("variables_adicionales_notas", "") or ""

    row = i + 1
    vals = [
        i,
        art.get("id", ""),
        art.get("titulo", ""),
        get_autores(art),
        art.get("year", ""),
        str(venue)[:80] if venue else "",
        art.get("relevancia_general", ""),
        get_granularidad(intenciones),
        len(intenciones),
        n_hall,
    ]
    set_row(ws1, row, vals, FILLS["row_even"], FILLS["row_odd"])
    ws1.row_dimensions[row].height = 32

set_col_widths(ws1, [5, 16, 55, 30, 6, 30, 14, 22, 14, 20])

# ════════════════════════════════════════════════════════════════════════════
# HOJA 2: MÉTODOS
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("2_Metodos")

h2 = ["ID Artículo", "Título (abrev.)", "Año", "Relevancia",
       "Métodos Generales", "Métodos Específicos", "Técnicas Metodológicas", "N Métodos Específicos"]
write_headers(ws2, h2, "header_teal")

for i, art in enumerate(arts, start=1):
    metodos_g = safe_list(art, "variables_principales", "metodos_general")
    metodos_e = safe_list(art, "variables_principales", "metodos_especifico")
    tecnicas  = safe_list(art, "variables_adicionales", "metodologica")

    vals = [
        art.get("id", ""),
        art.get("titulo", "")[:60],
        art.get("year", ""),
        art.get("relevancia_general", ""),
        "; ".join(metodos_g),
        "; ".join(metodos_e),
        "; ".join(tecnicas),
        len(metodos_e),
    ]
    set_row(ws2, i + 1, vals, FILLS["row_teal_e"], FILLS["row_odd"])

set_col_widths(ws2, [16, 45, 6, 12, 35, 55, 45, 14])

# ════════════════════════════════════════════════════════════════════════════
# HOJA 3: INTENCIONES COMUNICATIVAS
# ════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("3_Intenciones")

# Todas las intenciones únicas en el corpus
all_intenciones = set()
for art in arts:
    for i in safe_list(art, "variables_principales", "intenciones"):
        all_intenciones.add(i)
all_intenciones = sorted(all_intenciones)

h3 = ["ID Artículo", "Título (abrev.)", "Año", "Relevancia",
      "Tipo (Binario/Granular)", "N Dimensiones", "Intenciones (lista)"] + all_intenciones
write_headers(ws3, h3, "header_green")

for i, art in enumerate(arts, start=1):
    intenciones = safe_list(art, "variables_principales", "intenciones")
    # Columnas booleanas por intención
    bool_cols = ["Sí" if intent in intenciones else "" for intent in all_intenciones]

    vals = [
        art.get("id", ""),
        art.get("titulo", "")[:55],
        art.get("year", ""),
        art.get("relevancia_general", ""),
        get_granularidad(intenciones),
        len(intenciones),
        "; ".join(intenciones),
    ] + bool_cols
    set_row(ws3, i + 1, vals, FILLS["row_green_e"], FILLS["row_green_o"])

widths_base = [16, 42, 6, 12, 22, 14, 50]
widths_bool = [10] * len(all_intenciones)
set_col_widths(ws3, widths_base + widths_bool)

# ════════════════════════════════════════════════════════════════════════════
# HOJA 4: DATASETS
# ════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("4_Datasets")

h4 = ["ID Artículo", "Título (abrev.)", "Año", "Relevancia",
      "N Datasets", "Nombres Datasets", "Dataset 1 - Referencia", "Dataset 2 - Referencia"]
write_headers(ws4, h4, "header_purple")

for i, art in enumerate(arts, start=1):
    datasets     = safe_list(art, "variables_principales", "dataset")
    dataset_info = safe_list(art, "variables_principales", "dataset_info")

    ref1 = ref2 = ""
    if isinstance(dataset_info, list) and len(dataset_info) > 0:
        d = dataset_info[0]
        if isinstance(d, dict):
            ref1 = d.get("referencia", "") or d.get("nombre", "")
    if isinstance(dataset_info, list) and len(dataset_info) > 1:
        d = dataset_info[1]
        if isinstance(d, dict):
            ref2 = d.get("referencia", "") or d.get("nombre", "")

    vals = [
        art.get("id", ""),
        art.get("titulo", "")[:55],
        art.get("year", ""),
        art.get("relevancia_general", ""),
        len(datasets),
        "; ".join(datasets[:5]),
        str(ref1)[:200],
        str(ref2)[:200],
    ]
    set_row(ws4, i + 1, vals, FILLS["row_purple_e"], FILLS["row_odd"])

set_col_widths(ws4, [16, 42, 6, 12, 10, 50, 80, 80])

# ════════════════════════════════════════════════════════════════════════════
# HOJA 5: MÉTRICAS DE EVALUACIÓN
# ════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("5_Metricas")

all_metricas = set()
for art in arts:
    for m in safe_list(art, "variables_principales", "metricas"):
        all_metricas.add(m)
all_metricas = sorted(all_metricas)

h5 = ["ID Artículo", "Título (abrev.)", "Año", "Relevancia",
      "N Métricas", "Métricas (lista)"] + all_metricas
write_headers(ws5, h5, "header_brown")

for i, art in enumerate(arts, start=1):
    metricas = safe_list(art, "variables_principales", "metricas")
    bool_cols = ["Sí" if m in metricas else "" for m in all_metricas]

    vals = [
        art.get("id", ""),
        art.get("titulo", "")[:55],
        art.get("year", ""),
        art.get("relevancia_general", ""),
        len(metricas),
        "; ".join(metricas),
    ] + bool_cols
    set_row(ws5, i + 1, vals, FILLS["row_brown_e"], FILLS["row_odd"])

widths_b = [16, 42, 6, 12, 10, 45]
set_col_widths(ws5, widths_b + [14] * len(all_metricas))

# ════════════════════════════════════════════════════════════════════════════
# HOJA 6: CARACTERÍSTICAS LINGÜÍSTICAS Y PLATAFORMA
# ════════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("6_Contexto")

h6 = ["ID Artículo", "Título (abrev.)", "Año", "Relevancia",
      "Idioma(s)", "Plataforma(s)", "Período temporal", "Técnicas metodológicas"]
write_headers(ws6, h6, "header_gray")

for i, art in enumerate(arts, start=1):
    idiomas   = safe_list(art, "variables_adicionales", "linguistica")
    plataform = safe_list(art, "variables_adicionales", "plataforma")
    temporal  = art.get("clasificacion", {}).get("variables_adicionales", {}).get("temporal", "") or ""
    tecnicas  = safe_list(art, "variables_adicionales", "metodologica")

    vals = [
        art.get("id", ""),
        art.get("titulo", "")[:55],
        art.get("year", ""),
        art.get("relevancia_general", ""),
        "; ".join(idiomas),
        "; ".join(plataform),
        str(temporal),
        "; ".join(tecnicas[:4]),
    ]
    set_row(ws6, i + 1, vals, FILLS["row_gray_e"], FILLS["row_odd"])

set_col_widths(ws6, [16, 42, 6, 12, 20, 30, 14, 55])

# ════════════════════════════════════════════════════════════════════════════
# HOJA 7: HALLAZGOS EMERGENTES
# ════════════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("7_Hallazgos_Emergentes")

h7 = ["ID Artículo", "Título (abrev.)", "Año", "Relevancia",
      "N Hallazgos", "Hallazgo 1", "Hallazgo 2", "Hallazgo 3",
      "Hallazgo 4", "Hallazgo 5", "Hallazgos adicionales (concatenados)"]
write_headers(ws7, h7, "header_navy")

for i, art in enumerate(arts, start=1):
    hallazgos = art.get("hallazgos_nuevos", [])
    if not isinstance(hallazgos, list):
        hallazgos = []

    h = [str(x) for x in hallazgos]
    adicionales = "; ".join(h[5:]) if len(h) > 5 else ""

    vals = [
        art.get("id", ""),
        art.get("titulo", "")[:55],
        art.get("year", ""),
        art.get("relevancia_general", ""),
        len(h),
        h[0] if len(h) > 0 else "",
        h[1] if len(h) > 1 else "",
        h[2] if len(h) > 2 else "",
        h[3] if len(h) > 3 else "",
        h[4] if len(h) > 4 else "",
        adicionales,
    ]
    set_row(ws7, i + 1, vals, FILLS["row_even"], FILLS["row_odd"])
    ws7.row_dimensions[i + 1].height = 40

set_col_widths(ws7, [16, 42, 6, 12, 10, 55, 55, 55, 55, 55, 80])

# ── Guardar ───────────────────────────────────────────────────────────────────
output = f"{BASE}/prisma/DATA_EXTRACTION_FORM_GENERADO.xlsx"
wb.save(output)
print(f"\nArchivo guardado: {output}")
print(f"\nHojas generadas:")
print(f"  1_Indice              : {len(arts)} filas")
print(f"  2_Metodos             : {len(arts)} filas")
print(f"  3_Intenciones         : {len(arts)} filas | {len(all_intenciones)} dimensiones como columnas")
print(f"  4_Datasets            : {len(arts)} filas")
print(f"  5_Metricas            : {len(arts)} filas | {len(all_metricas)} métricas como columnas")
print(f"  6_Contexto            : {len(arts)} filas")
print(f"  7_Hallazgos_Emergentes: {len(arts)} filas")
